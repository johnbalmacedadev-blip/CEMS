"""
Export plates from (PRIVATE) AVAILABLE-RESERVED-RELEASED UNITS.xlsx
into storage/app/excel_units_snapshot.json for Unit Report reconciliation.
"""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

DEFAULT_SRC = Path(
    r"c:\xampp\htdocs\CarEmpire\NEW CAR EMPIRE FOLDER-20260727T012011Z-1-001\NEW CAR EMPIRE FOLDER\(PRIVATE) AVAILABLE-RESERVED-RELEASED UNITS.xlsx"
)
OUT = Path(
    r"c:\xampp\htdocs\CarEmpire\System V.1 - June 1 Updated\storage\app\excel_units_snapshot.json"
)
SRC = Path(__import__("os").environ.get("EXCEL_UNITS_SRC") or (DEFAULT_SRC if DEFAULT_SRC.exists() else DEFAULT_SRC))

# sheet name → designation meta
TAB_MAP = {
    "FLAGSHIP - UNITS": {
        "key": "flagship_available",
        "status": "Available",
        "branch": "Flagship",
        "label": "Flagship Available",
    },
    "FLAGSHIP - RESERVED": {
        "key": "flagship_reserved",
        "status": "Reserved",
        "branch": "Flagship",
        "label": "Flagship Reserved",
    },
    "FLAGSHIP - RELEASED": {
        "key": "flagship_released",
        "status": "Released",
        "branch": "Flagship",
        "label": "Flagship Released",
    },
    "ANNEX - UNITS": {
        "key": "annex_available",
        "status": "Available",
        "branch": "Annex",
        "label": "Annex Available",
    },
    "ANNEX - RESERVED": {
        "key": "annex_reserved",
        "status": "Reserved",
        "branch": "Annex",
        "label": "Annex Reserved",
    },
    "ANNEX RELEASED": {
        "key": "annex_released",
        "status": "Released",
        "branch": "Annex",
        "label": "Annex Released",
    },
    "FLAGSHIP- FORFEITREFUND": {
        "key": "flagship_forfeited",
        "status": "Forfeited",
        "branch": "Flagship",
        "label": "Flagship Forfeited",
    },
}


def norm_header(h) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).replace("\n", " ").upper()).strip()


def norm_plate(v) -> str:
    if v is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(v).upper())


def parse_year(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        y = int(v)
        return y if 1980 <= y <= 2100 else None
    s = str(v).strip()
    m = re.search(r"(19|20)\d{2}", s)
    return int(m.group(0)) if m else None


def to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def looks_like_plate(v) -> bool:
    if v is None:
        return False
    s = str(v).strip()
    if not s or len(s) > 15:
        return False
    # skip section headers / numbers-only line items that are costs
    pn = norm_plate(s)
    if len(pn) < 3:
        return False
    # must have at least one letter and one digit typically; allow pure numeric rare plates
    has_letter = any(c.isalpha() for c in pn)
    has_digit = any(c.isdigit() for c in pn)
    if not has_digit:
        return False
    # reject values that look like money/year only
    if pn.isdigit() and len(pn) > 6:
        return False
    if not has_letter and len(pn) <= 2:
        return False
    return True


def find_cols(ws, header_row: int = 1):
    headers = {}
    max_c = ws.max_column or 80
    for c in range(1, max_c + 1):
        h = norm_header(ws.cell(header_row, c).value)
        if h:
            headers[h] = c

    def find(*needles):
        for needle in needles:
            for h, c in headers.items():
                if needle in h:
                    return c
        return None

    plate = find("PLATE NUMBER", "PLATE")
    release = find("FORMATTED RELEASE DATE", "RELEASE DATE")
    sale = find("FORMATTED SALE DATE", "SALE DATE")
    return plate, release, sale, headers


def extract_sheet(ws, meta: dict) -> dict:
    plate_col, release_col, sale_col, headers = find_cols(ws, 1)

    # Forfeit sheet has different layout: plate often col 6
    if meta["status"] == "Forfeited" and not plate_col:
        # header row 1: PLATE NUMBER around col 6
        for c in range(1, 20):
            if "PLATE" in norm_header(ws.cell(1, c).value):
                plate_col = c
                break

    if not plate_col:
        # fallback common positions
        plate_col = 9 if meta["status"] != "Forfeited" else 6

    rows = []
    section = None
    max_r = ws.max_row or 0
    for r in range(2, max_r + 1):
        plate_raw = ws.cell(r, plate_col).value
        # section heading often in year/make columns
        b = ws.cell(r, 2).value
        if not looks_like_plate(plate_raw):
            if isinstance(b, str) and len(b.strip()) > 8:
                section = b.strip()
            elif isinstance(plate_raw, str) and len(plate_raw.strip()) > 8 and not looks_like_plate(plate_raw):
                section = plate_raw.strip()
            continue

        plate = norm_plate(plate_raw)
        release_date = to_date(ws.cell(r, release_col).value) if release_col else None
        sale_date = to_date(ws.cell(r, sale_col).value) if sale_col else None
        year_cell = ws.cell(r, 2).value if meta["status"] != "Forfeited" else ws.cell(r, 3).value
        year = parse_year(year_cell)
        make = ws.cell(r, 3).value if meta["status"] != "Forfeited" else ws.cell(r, 4).value
        model = ws.cell(r, 4).value if meta["status"] != "Forfeited" else ws.cell(r, 5).value

        rows.append(
            {
                "excel_row": r,
                "plate": plate,
                "plate_raw": str(plate_raw).strip(),
                "release_date": release_date,
                "sale_date": sale_date,
                "year": year,
                "year_raw": None if isinstance(year_cell, (int, float)) or year_cell is None else str(year_cell),
                "make": str(make).strip() if make else None,
                "model": str(model).strip() if model else None,
                "section": section,
            }
        )

    # Duplicate plate analysis (same plate multiple times with different release dates)
    by_plate: dict[str, list] = {}
    for row in rows:
        by_plate.setdefault(row["plate"], []).append(row)

    duplicates = []
    for plate, items in by_plate.items():
        if len(items) < 2:
            continue
        releases = [i["release_date"] for i in items if i.get("release_date")]
        duplicates.append(
            {
                "plate": plate,
                "occurrences": [
                    {
                        "excel_row": i["excel_row"],
                        "release_date": i["release_date"],
                        "sale_date": i["sale_date"],
                        "section": i["section"],
                    }
                    for i in items
                ],
                "oldest_release": min(releases) if releases else None,
                "newest_release": max(releases) if releases else None,
            }
        )

    return {
        **meta,
        "tab": None,  # filled by caller
        "excel_row_count": len(rows),
        "excel_unique_plates": len(by_plate),
        "duplicate_plate_count": len(duplicates),
        "duplicates": duplicates,
        "rows": rows,
    }


def main():
    if not SRC.exists():
        raise SystemExit(f"Excel not found: {SRC}")

    wb = openpyxl.load_workbook(SRC, data_only=True)
    designations = []
    for sheet_name, meta in TAB_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"SKIP missing sheet: {sheet_name}")
            continue
        print(f"Reading {sheet_name}...")
        data = extract_sheet(wb[sheet_name], dict(meta))
        data["tab"] = sheet_name
        designations.append(data)
        print(
            f"  rows={data['excel_row_count']} unique={data['excel_unique_plates']} dups={data['duplicate_plate_count']}"
        )

    payload = {
        "source": str(SRC),
        "source_name": SRC.name,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "designations": designations,
    }
    OUT.write_text(json.dumps(payload), encoding="utf-8")
    print(f"Wrote {OUT} ({OUT.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
